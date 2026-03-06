"""
知識庫匯出服務 - Excel 匯出功能

支援功能：
- Phase 1: 基礎 Excel 匯出（單工作表、基本格式）
- Phase 2: 進階格式化（多工作表、自動調整欄寬）
- Phase 3: 效能優化（分批處理、進度追蹤）

實作日期：2025-11-21
重構日期：2025-11-21 - 改用統一 Job 系統
"""

import os
import asyncio
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import asyncpg
from asyncpg.pool import Pool

# 引入統一 Job 服務
from services.unified_job_service import UnifiedJobService


class KnowledgeExportService(UnifiedJobService):
    """知識庫匯出服務（已整合到統一 Job 系統）"""

    def __init__(self, db_pool: Optional[Pool] = None):
        # 初始化父類（統一 Job 服務）
        super().__init__(db_pool)

        self.export_dir = Path('/tmp/knowledge_exports')
        self.export_dir.mkdir(exist_ok=True)

        # Phase 2: 進階格式化配置
        self.header_fill = PatternFill(
            start_color='4472C4',
            end_color='4472C4',
            fill_type='solid'
        )
        self.header_font = Font(color='FFFFFF', bold=True, size=11)
        self.header_alignment = Alignment(horizontal='center', vertical='center')

        # Phase 3: 效能配置
        self.batch_size = 1000  # 分批處理大小
        self.max_rows_per_sheet = 100000  # 單工作表最大行數

    @staticmethod
    def sanitize_for_excel(text) -> str:
        """
        清理文字以符合 Excel 格式要求

        移除控制字元和無效的 Unicode 字元
        處理列表和非字串類型
        """
        # 處理 None
        if text is None:
            return ''

        # 處理列表 - 轉換為字串
        if isinstance(text, (list, tuple)):
            text = ';'.join(str(x) for x in text)

        # 轉換為字串
        if not isinstance(text, str):
            text = str(text)

        # 移除 Excel 不允許的控制字元 (0x00-0x1F，除了 0x09, 0x0A, 0x0D)
        import re
        # 保留 tab (0x09), newline (0x0A), carriage return (0x0D)
        sanitized = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', text)

        # 限制單元格長度 (Excel 限制為 32,767 字元)
        if len(sanitized) > 32767:
            sanitized = sanitized[:32764] + "..."

        return sanitized

    # ==================== 標準匯出功能（匯出/匯入兼容 + 支援大量資料分批處理）====================

    async def export_knowledge_standard(
        self,
        knowledge_list: List[Dict],
        output_filename: str = None,
        progress_callback: callable = None
    ) -> str:
        """
        標準 Excel 匯出（與匯入格式完全兼容 + 支援大量資料分批處理）

        使用與 knowledge_import 相同的格式，確保匯出的資料可以直接匯入。
        支援分批處理大量資料（10萬+ 筆），避免記憶體溢出。

        標準欄位：
        - question_summary (問題摘要) - 必填
        - answer (答案) - 必填
        - scope (作用域) - global/vendor/customized
        - vendor_id (業者ID) - 可選
        - business_types (業態類型) - 逗號分隔
        - target_user (目標用戶) - 逗號分隔
        - intent_names (意圖名稱) - 逗號分隔
        - keywords (關鍵字) - 逗號分隔
        - priority (優先級) - 0 或 1

        Args:
            knowledge_list: 知識列表
            output_filename: 輸出檔名（可選）
            progress_callback: 進度回調函數（可選）

        Returns:
            str: 匯出檔案路徑
        """
        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"knowledge_export_{timestamp}.xlsx"

        output_path = self.export_dir / output_filename
        total_count = len(knowledge_list)

        print(f"📤 開始標準格式匯出...")
        print(f"   總筆數: {total_count}")
        if total_count > self.batch_size:
            batch_count = (total_count + self.batch_size - 1) // self.batch_size
            print(f"   批次大小: {self.batch_size}")
            print(f"   批次數量: {batch_count}")

        # 使用 openpyxl 建立工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "知識庫"

        # 標準欄位標題（與匯入格式完全一致，支援 ID 更新）
        headers = [
            'id',                # 知識ID（用於更新，新增時可留空）
            'question_summary',  # 問題摘要
            'answer',            # 答案
            'scope',             # 作用域
            'vendor_id',         # 業者ID
            'business_types',    # 業態類型
            'target_user',       # 目標用戶
            'intent_names',      # 意圖名稱
            'keywords',          # 關鍵字
            'priority'           # 優先級
        ]
        ws.append(headers)

        # 格式化標題行
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 分批寫入資料（支援大量資料）
        batch_count = (total_count + self.batch_size - 1) // self.batch_size

        for batch_num in range(batch_count):
            start_idx = batch_num * self.batch_size
            end_idx = min((batch_num + 1) * self.batch_size, total_count)
            batch = knowledge_list[start_idx:end_idx]

            # 寫入當前批次
            for knowledge in batch:
                # 處理陣列欄位：轉換為逗號分隔字串
                business_types = ','.join(knowledge.get('business_types', [])) if knowledge.get('business_types') else ''
                target_user_list = knowledge.get('target_user', [])
                if isinstance(target_user_list, str):
                    target_user = target_user_list
                else:
                    target_user = ','.join(target_user_list) if target_user_list else ''

                keywords = ','.join(knowledge.get('keywords', [])) if knowledge.get('keywords') else ''
                intent_names = knowledge.get('intent_name', '')  # 從 JOIN 取得的主要意圖

                ws.append([
                    knowledge.get('id', ''),  # 知識ID（用於更新）
                    self.sanitize_for_excel(knowledge.get('question_summary', '')),
                    self.sanitize_for_excel(knowledge.get('answer', '')),
                    knowledge.get('scope', 'global'),
                    knowledge.get('vendor_id') or '',  # 可能是 None
                    business_types,
                    target_user,
                    intent_names,
                    keywords,
                    knowledge.get('priority', 0)
                ])

            # 進度回調
            if batch_count > 1:  # 只在分批處理時顯示進度
                progress = int((end_idx / total_count) * 100)
                print(f"   ⏳ 進度: {end_idx}/{total_count} ({progress}%)")

                if progress_callback:
                    await progress_callback(progress, end_idx, total_count)

            # 讓出 CPU（避免阻塞）
            if batch_count > 1:
                await asyncio.sleep(0.01)

        # 調整欄寬
        column_widths = {
            'A': 10,  # id
            'B': 40,  # question_summary
            'C': 60,  # answer
            'D': 12,  # scope
            'E': 10,  # vendor_id
            'F': 25,  # business_types
            'G': 15,  # target_user
            'H': 20,  # intent_names
            'I': 20,  # keywords
            'J': 10   # priority
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 答案欄自動換行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 凍結首列
        ws.freeze_panes = 'A2'

        # 自動篩選
        ws.auto_filter.ref = ws.dimensions

        # 儲存檔案
        print(f"   💾 儲存檔案中...")
        wb.save(output_path)

        file_size = os.path.getsize(output_path) / 1024  # KB
        print(f"✅ 標準格式匯出完成")
        print(f"   檔案: {output_filename}")
        print(f"   大小: {file_size:.2f} KB")
        print(f"   ℹ️  此檔案可直接用於匯入功能")

        return str(output_path)

    # ==================== 資料庫查詢（輔助方法）====================

    async def get_knowledge_from_db(
        self,
        filters: Dict = None
    ) -> List[Dict]:
        """
        從資料庫查詢知識列表

        注意：總是匯出所有知識，不過濾 vendor_id

        Args:
            filters: 篩選條件（保留以兼容舊版，但 vendor_id 會被忽略）
                - intent_ids: 意圖 ID 列表
                - priority: 優先級（0/1/null）
                - is_active: 是否啟用

        Returns:
            List[Dict]: 知識列表
        """
        if not self.db_pool:
            raise Exception("資料庫連線池未初始化")

        filters = filters or {}

        query = """
            SELECT DISTINCT ON (kb.id)
                kb.id,
                kb.question_summary,
                kb.answer,
                kb.scope,
                kb.vendor_ids,
                kb.business_types,
                kb.target_user,
                kb.keywords,
                kb.priority,
                kb.source_type,
                kb.created_at,
                kb.updated_at,
                COALESCE(i.name, '') as intent_name,
                COALESCE(i.id, 0) as intent_id
            FROM knowledge_base kb
            LEFT JOIN knowledge_intent_mapping kim ON kb.id = kim.knowledge_id AND kim.intent_type = 'primary'
            LEFT JOIN intents i ON kim.intent_id = i.id
            WHERE 1=1
        """

        params = []
        param_idx = 1

        # 意圖篩選
        if filters.get('intent_ids'):
            query += f" AND i.id = ANY(${param_idx})"
            params.append(filters['intent_ids'])
            param_idx += 1

        # 優先級篩選
        if filters.get('priority') is not None:
            query += f" AND kb.priority = ${param_idx}"
            params.append(filters['priority'])
            param_idx += 1

        # 啟用狀態篩選
        if filters.get('is_active') is not None:
            query += f" AND kb.is_active = ${param_idx}"
            params.append(filters['is_active'])
            param_idx += 1

        # vendor_ids 篩選
        if filters.get('vendor_id') is not None:
            query += f" AND (array_length(kb.vendor_ids, 1) IS NULL OR kb.vendor_ids && ARRAY[${param_idx}]::int[])"
            params.append(filters['vendor_id'])
            param_idx += 1
        elif filters.get('vendor_id') is None and 'vendor_id' in filters:
            # 明確指定 vendor_id=None 時，只匯出通用知識
            query += " AND array_length(kb.vendor_ids, 1) IS NULL"

        # DISTINCT ON 要求 ORDER BY 的第一個欄位必須是 DISTINCT ON 的欄位
        query += " ORDER BY kb.id ASC, kim.id ASC"

        async with self.db_pool.acquire() as conn:
            rows = await conn.fetch(query, *params)

            knowledge_list = []
            for row in rows:
                knowledge_list.append({
                    'id': row['id'],
                    'question_summary': row['question_summary'],
                    'answer': row['answer'],
                    'scope': row['scope'],
                    'vendor_id': row['vendor_id'],
                    'keywords': row['keywords'] or [],
                    'business_types': row['business_types'] or [],
                    'target_user': row['target_user'],
                    'priority': row['priority'],
                    'source_type': row['source_type'],
                    'created_at': row['created_at'],
                    'updated_at': row['updated_at'],
                    'intent_name': row['intent_name'],
                    'intent_id': row['intent_id']
                })

            return knowledge_list

    async def get_intents_from_db(self) -> List[Dict]:
        """從資料庫查詢意圖列表"""
        if not self.db_pool:
            raise Exception("資料庫連線池未初始化")

        query = """
            SELECT id, name, description
            FROM intents
            WHERE is_enabled = TRUE
            ORDER BY id
        """

        async with self.db_pool.acquire() as conn:
            rows = await conn.fetch(query)

            intents = []
            for row in rows:
                intents.append({
                    'id': row['id'],
                    'name': row['name'],
                    'description': row['description']
                })

            return intents

    # ==================== 背景任務處理 ====================

    async def process_export_job(
        self,
        job_id: str,
        vendor_id: Optional[int],
        export_mode: str,
        include_intents: bool,
        include_metadata: bool,
        user_id: str
    ):
        """
        背景任務：處理匯出作業

        Args:
            job_id: 作業 ID
            vendor_id: 業者 ID（可選）
            export_mode: 匯出模式（basic, formatted, optimized）
            include_intents: 是否包含意圖對照表
            include_metadata: 是否包含匯出資訊
            user_id: 使用者 ID
        """
        import uuid
        import json

        print(f"\n{'='*60}")
        print(f"🔄 開始處理匯出作業")
        print(f"   Job ID: {job_id}")
        print(f"   業者 ID: {vendor_id or '通用知識'}")
        print(f"   匯出模式: {export_mode}")
        print(f"{'='*60}\n")

        try:
            # 1. 更新狀態為 processing
            await self.update_status(
                job_id,
                status='processing',
                progress={'stage': 'fetching_data', 'message': '正在從資料庫查詢資料...'}
            )

            # 2. 從資料庫查詢知識
            print(f"📊 查詢知識資料...")
            # 根據 vendor_id 過濾知識
            # 當 vendor_id=None 時，匯出所有知識（不過濾）
            # 當 vendor_id=數字 時，只匯出該業者的知識
            filters = {} if vendor_id is None else {'vendor_id': vendor_id}
            knowledge_list = await self.get_knowledge_from_db(filters=filters)
            total_count = len(knowledge_list)

            if total_count == 0:
                raise Exception("沒有可匯出的知識資料")

            print(f"✅ 查詢完成，共 {total_count} 筆知識")

            # 3. 查詢意圖對照表（如果需要）
            intents = []
            if include_intents:
                print(f"📊 查詢意圖對照表...")
                intents = await self.get_intents_from_db()
                print(f"✅ 查詢完成，共 {len(intents)} 個意圖")

            # 4. 準備匯出資訊（如果需要）
            export_info = {}
            if include_metadata:
                export_info = {
                    '匯出時間': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '匯出人員': user_id,
                    '業者 ID': vendor_id or '通用知識',
                    '匯出模式': export_mode,
                    '知識總數': total_count,
                    '意圖總數': len(intents)
                }

            # 5. 更新進度
            await self.update_status(
                job_id,
                status='processing',
                progress={
                    'stage': 'exporting',
                    'message': f'正在匯出 {total_count} 筆資料...',
                    'current': 0,
                    'total': total_count
                },
                total_records=total_count
            )

            # 6. 使用標準匯出方法（與匯入格式完全兼容）
            # 注意：無論 export_mode 是什麼，統一使用標準格式
            output_filename = f"export_{job_id}.xlsx"

            print(f"📤 使用標準匯出格式（可直接匯入）...")
            print(f"   ℹ️  所有匯出統一使用標準格式，確保匯入兼容性")
            file_path = await self.export_knowledge_standard(
                knowledge_list=knowledge_list,
                output_filename=output_filename
            )

            # 7. 取得檔案大小
            file_size = os.path.getsize(file_path)

            print(f"✅ 匯出完成")
            print(f"   檔案路徑: {file_path}")
            print(f"   檔案大小: {file_size / 1024:.2f} KB")

            # 8. 更新為完成狀態（使用統一 Job 服務的方法）
            await self.update_status(
                job_id,
                status='completed',
                result={
                    'exported': total_count,
                    'file_path': str(file_path),
                    'file_size_kb': round(file_size / 1024, 2),
                    'file_size_bytes': file_size,
                    'export_mode': export_mode,
                    'vendor_id': vendor_id
                },
                success_records=total_count,
                file_path=str(file_path),
                file_size_bytes=file_size
            )

            print(f"{'='*60}")
            print(f"✅ 匯出作業完成 (Job ID: {job_id})")
            print(f"{'='*60}\n")

        except Exception as e:
            import traceback
            error_message = str(e)
            print(f"❌ 完整錯誤堆棧:")
            traceback.print_exc()
            print(f"❌ 匯出作業失敗: {error_message}")

            # 更新為失敗狀態（使用統一 Job 服務的方法）
            await self.update_status(
                job_id,
                status='failed',
                error_message=error_message
            )

            print(f"{'='*60}")
            print(f"❌ 匯出作業失敗 (Job ID: {job_id})")
            print(f"{'='*60}\n")

    # ✅ _update_job_status 方法已移除，改用父類 UnifiedJobService 的 update_status() 方法
